"""
THERESE v2 - CRM Export Service

Multi-format export for CRM data (CSV, Excel, JSON).
Part of the "Local First" architecture.
"""

import csv
import io
import json
import logging
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any, Literal

from app.models.entities import Contact, Deliverable, Project
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

logger = logging.getLogger(__name__)


ExportFormat = Literal["csv", "xlsx", "json"]


@dataclass
class ExportResult:
    """Result of an export operation."""
    data: bytes
    filename: str
    content_type: str
    row_count: int


# ============================================================
# Export Configuration
# ============================================================

# Column mappings for exports
CONTACT_COLUMNS = [
    ("id", "ID"),
    ("first_name", "Prenom"),
    ("last_name", "Nom"),
    ("company", "Entreprise"),
    ("email", "Email"),
    ("phone", "Telephone"),
    ("stage", "Stage"),
    ("score", "Score"),
    ("source", "Source"),
    ("tags", "Tags"),
    ("notes", "Notes"),
    ("last_interaction", "Derniere interaction"),
    ("created_at", "Date creation"),
    ("updated_at", "Date modification"),
]

PROJECT_COLUMNS = [
    ("id", "ID"),
    ("name", "Nom"),
    ("description", "Description"),
    ("contact_id", "ID Contact"),
    ("status", "Statut"),
    ("budget", "Budget"),
    ("notes", "Notes"),
    ("tags", "Tags"),
    ("created_at", "Date creation"),
    ("updated_at", "Date modification"),
]

DELIVERABLE_COLUMNS = [
    ("id", "ID"),
    ("project_id", "ID Projet"),
    ("title", "Titre"),
    ("description", "Description"),
    ("status", "Statut"),
    ("due_date", "Date echeance"),
    ("completed_at", "Date completion"),
    ("created_at", "Date creation"),
    ("updated_at", "Date modification"),
]


# ============================================================
# Helper Functions
# ============================================================


def _format_value(value: Any) -> str:
    """Format a value for export."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _parse_tags(tags_json: str | None) -> str:
    """Parse JSON tags to comma-separated string."""
    if not tags_json:
        return ""
    try:
        tags = json.loads(tags_json)
        if isinstance(tags, list):
            return ", ".join(str(t).strip() for t in tags if t)
        return str(tags)
    except (json.JSONDecodeError, TypeError):
        return tags_json


def _entity_to_row(entity: Any, columns: list[tuple[str, str]]) -> dict[str, Any]:
    """Convert entity to row dict with formatted values."""
    row = {}
    for attr, header in columns:
        value = getattr(entity, attr, None)
        if attr == "tags":
            value = _parse_tags(value)
        row[header] = _format_value(value)
    return row


# ============================================================
# CSV Export
# ============================================================


def export_to_csv(entities: list[Any], columns: list[tuple[str, str]]) -> bytes:
    """
    Export entities to CSV format.

    Args:
        entities: List of SQLModel entities
        columns: List of (attribute, header) tuples

    Returns:
        CSV content as bytes (UTF-8 with BOM for Excel compatibility)
    """
    output = io.StringIO()
    headers = [col[1] for col in columns]

    writer = csv.DictWriter(output, fieldnames=headers, quoting=csv.QUOTE_ALL)
    writer.writeheader()

    for entity in entities:
        row = _entity_to_row(entity, columns)
        writer.writerow(row)

    # Add BOM for Excel UTF-8 compatibility
    csv_content = output.getvalue()
    return ("\ufeff" + csv_content).encode("utf-8")


# ============================================================
# Excel Export
# ============================================================


def export_to_xlsx(
    entities: list[Any],
    columns: list[tuple[str, str]],
    sheet_name: str = "Export",
) -> bytes:
    """
    Export entities to Excel format with styling.

    Args:
        entities: List of SQLModel entities
        columns: List of (attribute, header) tuples
        sheet_name: Name of the Excel sheet

    Returns:
        Excel content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2451FF", end_color="2451FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [col[1] for col in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, entity in enumerate(entities, 2):
        row_data = _entity_to_row(entity, columns)
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        for row_idx in range(2, len(entities) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ============================================================
# JSON Export
# ============================================================


def export_to_json(entities: list[Any], columns: list[tuple[str, str]]) -> bytes:
    """
    Export entities to JSON format.

    Args:
        entities: List of SQLModel entities
        columns: List of (attribute, header) tuples

    Returns:
        JSON content as bytes
    """
    data = []
    for entity in entities:
        row = {}
        for attr, _ in columns:
            value = getattr(entity, attr, None)
            if attr == "tags" and value:
                try:
                    value = json.loads(value)
                except (json.JSONDecodeError, TypeError):
                    pass
            if isinstance(value, datetime):
                value = value.isoformat()
            row[attr] = value
        data.append(row)

    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


# ============================================================
# Export Service
# ============================================================


class CRMExportService:
    """
    Service for exporting CRM data in multiple formats.

    Supports CSV, Excel (XLSX), and JSON exports for:
    - Contacts
    - Projects
    - Deliverables
    - Full CRM export (all entities)
    """

    def __init__(self, session: AsyncSession):
        """
        Initialize export service.

        Args:
            session: AsyncSession for database access
        """
        self.session = session

    async def export_contacts(
        self,
        format: ExportFormat = "csv",
        stage: str | None = None,
        source: str | None = None,
    ) -> ExportResult:
        """
        Export contacts to specified format.

        Args:
            format: Export format (csv, xlsx, json)
            stage: Filter by pipeline stage
            source: Filter by contact source

        Returns:
            ExportResult with data and metadata
        """
        # Build query
        statement = select(Contact)
        if stage:
            statement = statement.where(Contact.stage == stage)
        if source:
            statement = statement.where(Contact.source == source)
        statement = statement.order_by(Contact.created_at.desc())

        result = await self.session.execute(statement)
        contacts = result.scalars().all()

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        if format == "csv":
            data = export_to_csv(contacts, CONTACT_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"contacts_{timestamp}.csv",
                content_type="text/csv; charset=utf-8",
                row_count=len(contacts),
            )
        elif format == "xlsx":
            data = export_to_xlsx(contacts, CONTACT_COLUMNS, "Contacts")
            return ExportResult(
                data=data,
                filename=f"contacts_{timestamp}.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                row_count=len(contacts),
            )
        else:  # json
            data = export_to_json(contacts, CONTACT_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"contacts_{timestamp}.json",
                content_type="application/json",
                row_count=len(contacts),
            )

    async def export_projects(
        self,
        format: ExportFormat = "csv",
        status: str | None = None,
        contact_id: str | None = None,
    ) -> ExportResult:
        """
        Export projects to specified format.

        Args:
            format: Export format (csv, xlsx, json)
            status: Filter by project status
            contact_id: Filter by linked contact

        Returns:
            ExportResult with data and metadata
        """
        statement = select(Project)
        if status:
            statement = statement.where(Project.status == status)
        if contact_id:
            statement = statement.where(Project.contact_id == contact_id)
        statement = statement.order_by(Project.created_at.desc())

        result = await self.session.execute(statement)
        projects = result.scalars().all()

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        if format == "csv":
            data = export_to_csv(projects, PROJECT_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"projects_{timestamp}.csv",
                content_type="text/csv; charset=utf-8",
                row_count=len(projects),
            )
        elif format == "xlsx":
            data = export_to_xlsx(projects, PROJECT_COLUMNS, "Projets")
            return ExportResult(
                data=data,
                filename=f"projects_{timestamp}.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                row_count=len(projects),
            )
        else:  # json
            data = export_to_json(projects, PROJECT_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"projects_{timestamp}.json",
                content_type="application/json",
                row_count=len(projects),
            )

    async def export_deliverables(
        self,
        format: ExportFormat = "csv",
        status: str | None = None,
        project_id: str | None = None,
    ) -> ExportResult:
        """
        Export deliverables to specified format.

        Args:
            format: Export format (csv, xlsx, json)
            status: Filter by deliverable status
            project_id: Filter by linked project

        Returns:
            ExportResult with data and metadata
        """
        statement = select(Deliverable)
        if status:
            statement = statement.where(Deliverable.status == status)
        if project_id:
            statement = statement.where(Deliverable.project_id == project_id)
        statement = statement.order_by(Deliverable.created_at.desc())

        result = await self.session.execute(statement)
        deliverables = result.scalars().all()

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        if format == "csv":
            data = export_to_csv(deliverables, DELIVERABLE_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"deliverables_{timestamp}.csv",
                content_type="text/csv; charset=utf-8",
                row_count=len(deliverables),
            )
        elif format == "xlsx":
            data = export_to_xlsx(deliverables, DELIVERABLE_COLUMNS, "Livrables")
            return ExportResult(
                data=data,
                filename=f"deliverables_{timestamp}.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                row_count=len(deliverables),
            )
        else:  # json
            data = export_to_json(deliverables, DELIVERABLE_COLUMNS)
            return ExportResult(
                data=data,
                filename=f"deliverables_{timestamp}.json",
                content_type="application/json",
                row_count=len(deliverables),
            )

    async def export_all(self, format: ExportFormat = "xlsx") -> ExportResult:
        """
        Export all CRM data to a single file.

        For Excel: Creates multiple sheets (Contacts, Projets, Livrables)
        For CSV/JSON: Creates a combined structure

        Args:
            format: Export format (csv, xlsx, json)

        Returns:
            ExportResult with data and metadata
        """
        # Fetch all data
        contacts_result = await self.session.execute(
            select(Contact).order_by(Contact.created_at.desc())
        )
        contacts = contacts_result.scalars().all()

        projects_result = await self.session.execute(
            select(Project).order_by(Project.created_at.desc())
        )
        projects = projects_result.scalars().all()

        deliverables_result = await self.session.execute(
            select(Deliverable).order_by(Deliverable.created_at.desc())
        )
        deliverables = deliverables_result.scalars().all()

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        total_count = len(contacts) + len(projects) + len(deliverables)

        if format == "xlsx":
            # Create workbook with multiple sheets
            wb = Workbook()

            # Contacts sheet
            ws_contacts = wb.active
            ws_contacts.title = "Contacts"
            _populate_xlsx_sheet(ws_contacts, contacts, CONTACT_COLUMNS)

            # Projects sheet
            ws_projects = wb.create_sheet("Projets")
            _populate_xlsx_sheet(ws_projects, projects, PROJECT_COLUMNS)

            # Deliverables sheet
            ws_deliverables = wb.create_sheet("Livrables")
            _populate_xlsx_sheet(ws_deliverables, deliverables, DELIVERABLE_COLUMNS)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return ExportResult(
                data=output.read(),
                filename=f"crm_export_{timestamp}.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                row_count=total_count,
            )

        elif format == "json":
            data = {
                "export_date": datetime.now(UTC).isoformat(),
                "contacts": [
                    {attr: getattr(c, attr, None) for attr, _ in CONTACT_COLUMNS}
                    for c in contacts
                ],
                "projects": [
                    {attr: getattr(p, attr, None) for attr, _ in PROJECT_COLUMNS}
                    for p in projects
                ],
                "deliverables": [
                    {attr: getattr(d, attr, None) for attr, _ in DELIVERABLE_COLUMNS}
                    for d in deliverables
                ],
            }

            # Serialize datetime objects
            def serialize(obj):
                if isinstance(obj, datetime):
                    return obj.isoformat()
                return obj

            json_str = json.dumps(data, default=serialize, ensure_ascii=False, indent=2)

            return ExportResult(
                data=json_str.encode("utf-8"),
                filename=f"crm_export_{timestamp}.json",
                content_type="application/json",
                row_count=total_count,
            )

        else:  # csv - combine all with type column
            output = io.StringIO()

            # Combined CSV with entity type
            all_columns = [("_type", "Type")] + CONTACT_COLUMNS
            headers = [col[1] for col in all_columns]

            writer = csv.DictWriter(output, fieldnames=headers, quoting=csv.QUOTE_ALL)
            writer.writeheader()

            for contact in contacts:
                row = _entity_to_row(contact, CONTACT_COLUMNS)
                row["Type"] = "Contact"
                writer.writerow(row)

            csv_content = output.getvalue()

            return ExportResult(
                data=("\ufeff" + csv_content).encode("utf-8"),
                filename=f"crm_export_{timestamp}.csv",
                content_type="text/csv; charset=utf-8",
                row_count=total_count,
            )


def _populate_xlsx_sheet(ws, entities: list[Any], columns: list[tuple[str, str]]):
    """Populate an Excel sheet with data and styling."""
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2451FF", end_color="2451FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [col[1] for col in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, entity in enumerate(entities, 2):
        row_data = _entity_to_row(entity, columns)
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        for row_idx in range(2, len(entities) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"
