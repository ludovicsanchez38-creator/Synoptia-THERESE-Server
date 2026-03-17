"""
THERESE v2 - CRM Import Service

Multi-format import for CRM data (CSV, Excel, JSON).
Part of the "Local First" architecture.
"""

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any, Literal

from app.models.entities import Contact, Deliverable, Project, generate_uuid
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

logger = logging.getLogger(__name__)


ImportFormat = Literal["csv", "xlsx", "json"]


@dataclass
class ImportError:
    """Error encountered during import."""
    row: int
    column: str | None
    message: str
    data: dict | None = None


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[ImportError] = field(default_factory=list)
    total_rows: int = 0

    @property
    def message(self) -> str:
        """Human-readable summary."""
        if self.success:
            return f"Import termine: {self.created} crees, {self.updated} mis a jour, {self.skipped} ignores"
        return f"Import avec erreurs: {len(self.errors)} erreurs sur {self.total_rows} lignes"


@dataclass
class ImportPreview:
    """Preview of import before execution."""
    total_rows: int
    sample_rows: list[dict]
    detected_columns: list[str]
    column_mapping: dict[str, str]
    validation_errors: list[ImportError]
    can_import: bool


# ============================================================
# Column Mappings
# ============================================================

# Flexible column name mappings (source -> internal)
CONTACT_COLUMN_MAPPING = {
    # ID
    "id": "id",
    "ID": "id",
    "identifiant": "id",
    # First name
    "first_name": "first_name",
    "prenom": "first_name",
    "Prenom": "first_name",
    "prénom": "first_name",
    "Prénom": "first_name",
    "firstname": "first_name",
    "FirstName": "first_name",
    # Last name
    "last_name": "last_name",
    "nom": "last_name",
    "Nom": "last_name",
    "lastname": "last_name",
    "LastName": "last_name",
    # Company
    "company": "company",
    "entreprise": "company",
    "Entreprise": "company",
    "societe": "company",
    "Societe": "company",
    "société": "company",
    "Société": "company",
    # Email
    "email": "email",
    "Email": "email",
    "e-mail": "email",
    "E-mail": "email",
    "courriel": "email",
    # Phone
    "phone": "phone",
    "telephone": "phone",
    "Telephone": "phone",
    "téléphone": "phone",
    "Téléphone": "phone",
    "tel": "phone",
    "Tel": "phone",
    # Stage
    "stage": "stage",
    "Stage": "stage",
    "etape": "stage",
    "Etape": "stage",
    # Score
    "score": "score",
    "Score": "score",
    # Source
    "source": "source",
    "Source": "source",
    "origine": "source",
    # Tags
    "tags": "tags",
    "Tags": "tags",
    "etiquettes": "tags",
    # Notes
    "notes": "notes",
    "Notes": "notes",
    "commentaires": "notes",
}

PROJECT_COLUMN_MAPPING = {
    # ID
    "id": "id",
    "ID": "id",
    # Name
    "name": "name",
    "Name": "name",
    "nom": "name",
    "Nom": "name",
    "titre": "name",
    "Titre": "name",
    # Description
    "description": "description",
    "Description": "description",
    # Contact ID
    "contact_id": "contact_id",
    "ContactID": "contact_id",
    "ClientID": "contact_id",
    "client_id": "contact_id",
    "ID Contact": "contact_id",
    # Status
    "status": "status",
    "Status": "status",
    "statut": "status",
    "Statut": "status",
    # Budget
    "budget": "budget",
    "Budget": "budget",
    # Notes
    "notes": "notes",
    "Notes": "notes",
    # Tags
    "tags": "tags",
    "Tags": "tags",
}

DELIVERABLE_COLUMN_MAPPING = {
    # ID
    "id": "id",
    "ID": "id",
    # Project ID
    "project_id": "project_id",
    "ProjectID": "project_id",
    "ID Projet": "project_id",
    # Title
    "title": "title",
    "Title": "title",
    "titre": "title",
    "Titre": "title",
    "nom": "title",
    "Nom": "title",
    # Description
    "description": "description",
    "Description": "description",
    # Status
    "status": "status",
    "Status": "status",
    "statut": "status",
    "Statut": "status",
    # Due date
    "due_date": "due_date",
    "DueDate": "due_date",
    "Date echeance": "due_date",
    "echeance": "due_date",
}


# ============================================================
# Parsing Helpers
# ============================================================


def _detect_format(content: bytes, filename: str | None = None) -> ImportFormat:
    """Detect file format from content or filename."""
    if filename:
        if filename.endswith(".csv"):
            return "csv"
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            return "xlsx"
        elif filename.endswith(".json"):
            return "json"

    # Try to detect from content
    try:
        content.decode("utf-8")
        # Check if JSON
        stripped = content.strip()
        if stripped.startswith(b"[") or stripped.startswith(b"{"):
            return "json"
        return "csv"
    except UnicodeDecodeError:
        return "xlsx"


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content to list of dicts."""
    # Try different encodings
    for encoding in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Impossible de decoder le fichier CSV")

    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _parse_xlsx(content: bytes, sheet_name: str | None = None) -> list[dict]:
    """Parse Excel content to list of dicts."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []

    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = cell
        data.append(row_dict)

    return data


def _parse_json(content: bytes) -> list[dict]:
    """Parse JSON content to list of dicts."""
    data = json.loads(content.decode("utf-8"))
    if isinstance(data, dict):
        # Check for nested structure
        if "contacts" in data:
            return data["contacts"]
        elif "projects" in data:
            return data["projects"]
        elif "deliverables" in data:
            return data["deliverables"]
        return [data]
    return data


# Field length limits (SEC-017)
FIELD_MAX_LENGTHS: dict[str, int] = {
    "first_name": 200,
    "last_name": 200,
    "company": 300,
    "email": 320,
    "phone": 50,
    "stage": 50,
    "source": 200,
    "tags": 1000,
    "notes": 5000,
    "name": 500,
    "title": 500,
    "description": 5000,
    "status": 50,
}

# Characters that trigger formula injection in spreadsheets (SEC-017)
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanitize_field(value: Any, field_name: str | None = None) -> Any:
    """
    Sanitize a field value for safe storage (SEC-017).

    - Strip whitespace
    - Enforce length limits
    - Neutralize formula injection prefixes (CSV injection defense)
    - Remove null bytes
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    # Strip and remove null bytes
    value = value.strip().replace("\x00", "")
    if not value:
        return None
    # Neutralize formula injection - prefix with single quote if dangerous
    if value and value[0] in FORMULA_PREFIXES:
        value = "'" + value
    # Enforce length limit
    if field_name and field_name in FIELD_MAX_LENGTHS:
        max_len = FIELD_MAX_LENGTHS[field_name]
        if len(value) > max_len:
            value = value[:max_len]
    return value


def _map_columns(row: dict, mapping: dict[str, str]) -> dict[str, Any]:
    """Map source columns to internal column names and sanitize values (SEC-017)."""
    result = {}
    for source_col, value in row.items():
        internal_col = mapping.get(source_col, source_col)
        if internal_col in mapping.values():
            result[internal_col] = _sanitize_field(value, internal_col)
    return result


def _parse_value(value: Any, field_type: str) -> Any:
    """Parse and validate a value based on field type."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if field_type == "int":
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None

    if field_type == "float":
        try:
            return float(str(value).strip().replace(",", "."))
        except (ValueError, TypeError):
            return None

    if field_type == "datetime":
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            for fmt in [
                "%Y-%m-%dT%H:%M:%S.%fZ",
                "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%d-%m-%Y",
            ]:
                try:
                    return datetime.strptime(value.strip(), fmt)
                except ValueError:
                    continue
        return None

    if field_type == "tags":
        if isinstance(value, list):
            return json.dumps(value)
        if isinstance(value, str):
            tags = [t.strip() for t in value.split(",") if t.strip()]
            return json.dumps(tags) if tags else None
        return None

    # String
    return str(value).strip() if value else None


def _validate_contact(data: dict) -> list[str]:
    """Validate contact data, return list of errors."""
    errors = []
    if not data.get("first_name") and not data.get("last_name") and not data.get("company"):
        errors.append("Au moins un nom ou une entreprise est requis")
    if data.get("email") and "@" not in str(data.get("email", "")):
        errors.append("Email invalide")
    return errors


def _validate_project(data: dict) -> list[str]:
    """Validate project data, return list of errors."""
    errors = []
    if not data.get("name"):
        errors.append("Le nom du projet est requis")
    return errors


def _validate_deliverable(data: dict) -> list[str]:
    """Validate deliverable data, return list of errors."""
    errors = []
    if not data.get("title"):
        errors.append("Le titre du livrable est requis")
    if not data.get("project_id"):
        errors.append("L'ID du projet est requis")
    return errors


# ============================================================
# Import Service
# ============================================================


class CRMImportService:
    """
    Service for importing CRM data from multiple formats.

    Supports CSV, Excel (XLSX), and JSON imports for:
    - Contacts
    - Projects
    - Deliverables
    """

    def __init__(self, session: AsyncSession):
        """
        Initialize import service.

        Args:
            session: AsyncSession for database access
        """
        self.session = session

    async def preview_contacts(
        self,
        content: bytes,
        filename: str | None = None,
        custom_mapping: dict[str, str] | None = None,
    ) -> ImportPreview:
        """
        Preview contact import without executing.

        Args:
            content: File content as bytes
            filename: Original filename for format detection
            custom_mapping: Custom column mapping to override defaults

        Returns:
            ImportPreview with sample data and validation results
        """
        format_type = _detect_format(content, filename)
        mapping = {**CONTACT_COLUMN_MAPPING, **(custom_mapping or {})}

        try:
            if format_type == "csv":
                raw_data = _parse_csv(content)
            elif format_type == "xlsx":
                raw_data = _parse_xlsx(content)
            else:
                raw_data = _parse_json(content)
        except Exception as e:
            return ImportPreview(
                total_rows=0,
                sample_rows=[],
                detected_columns=[],
                column_mapping={},
                validation_errors=[ImportError(row=0, column=None, message=str(e))],
                can_import=False,
            )

        if not raw_data:
            return ImportPreview(
                total_rows=0,
                sample_rows=[],
                detected_columns=[],
                column_mapping={},
                validation_errors=[ImportError(row=0, column=None, message="Aucune donnee trouvee")],
                can_import=False,
            )

        detected_columns = list(raw_data[0].keys()) if raw_data else []
        used_mapping = {col: mapping.get(col, col) for col in detected_columns if col in mapping}

        validation_errors = []
        sample_rows = []

        for idx, row in enumerate(raw_data[:5]):
            mapped = _map_columns(row, mapping)
            sample_rows.append(mapped)

            errors = _validate_contact(mapped)
            for error in errors:
                validation_errors.append(ImportError(row=idx + 1, column=None, message=error, data=mapped))

        can_import = len(validation_errors) == 0 or all(
            err.row > 5 for err in validation_errors
        )

        return ImportPreview(
            total_rows=len(raw_data),
            sample_rows=sample_rows,
            detected_columns=detected_columns,
            column_mapping=used_mapping,
            validation_errors=validation_errors[:10],
            can_import=can_import,
        )

    async def import_contacts(
        self,
        content: bytes,
        filename: str | None = None,
        custom_mapping: dict[str, str] | None = None,
        update_existing: bool = True,
    ) -> ImportResult:
        """
        Import contacts from file.

        Args:
            content: File content as bytes
            filename: Original filename for format detection
            custom_mapping: Custom column mapping to override defaults
            update_existing: Whether to update existing contacts by ID

        Returns:
            ImportResult with counts and errors
        """
        format_type = _detect_format(content, filename)
        mapping = {**CONTACT_COLUMN_MAPPING, **(custom_mapping or {})}

        try:
            if format_type == "csv":
                raw_data = _parse_csv(content)
            elif format_type == "xlsx":
                raw_data = _parse_xlsx(content)
            else:
                raw_data = _parse_json(content)
        except Exception as e:
            return ImportResult(
                success=False,
                errors=[ImportError(row=0, column=None, message=str(e))],
            )

        result = ImportResult(success=True, total_rows=len(raw_data))

        for idx, row in enumerate(raw_data):
            try:
                mapped = _map_columns(row, mapping)

                # Validate
                errors = _validate_contact(mapped)
                if errors:
                    result.errors.append(ImportError(
                        row=idx + 1,
                        column=None,
                        message="; ".join(errors),
                        data=mapped,
                    ))
                    result.skipped += 1
                    continue

                # Check for existing contact
                contact_id = mapped.get("id")
                existing = None

                if contact_id:
                    stmt = select(Contact).where(Contact.id == contact_id)
                    db_result = await self.session.execute(stmt)
                    existing = db_result.scalar_one_or_none()

                if existing and update_existing:
                    # Update existing
                    if mapped.get("first_name"):
                        existing.first_name = mapped["first_name"]
                    if mapped.get("last_name"):
                        existing.last_name = mapped["last_name"]
                    if mapped.get("company"):
                        existing.company = mapped["company"]
                    if mapped.get("email"):
                        existing.email = mapped["email"]
                    if mapped.get("phone"):
                        existing.phone = mapped["phone"]
                    if mapped.get("stage"):
                        existing.stage = mapped["stage"]
                    if mapped.get("score"):
                        existing.score = _parse_value(mapped["score"], "int") or existing.score
                    if mapped.get("source"):
                        existing.source = mapped["source"]
                    if mapped.get("tags"):
                        existing.tags = _parse_value(mapped["tags"], "tags")
                    if mapped.get("notes"):
                        existing.notes = mapped["notes"]

                    existing.updated_at = datetime.now(UTC)
                    self.session.add(existing)
                    result.updated += 1

                elif existing and not update_existing:
                    result.skipped += 1

                else:
                    # Create new
                    contact = Contact(
                        id=contact_id or generate_uuid(),
                        first_name=mapped.get("first_name"),
                        last_name=mapped.get("last_name"),
                        company=mapped.get("company"),
                        email=mapped.get("email"),
                        phone=mapped.get("phone"),
                        stage=mapped.get("stage", "contact"),
                        score=_parse_value(mapped.get("score"), "int") or 50,
                        source=mapped.get("source"),
                        tags=_parse_value(mapped.get("tags"), "tags"),
                        notes=mapped.get("notes"),
                        scope="global",
                    )
                    self.session.add(contact)
                    result.created += 1

            except Exception as e:
                logger.error(f"Error importing contact row {idx + 1}: {e}")
                result.errors.append(ImportError(row=idx + 1, column=None, message=str(e)))
                result.skipped += 1

        await self.session.commit()

        result.success = len(result.errors) == 0
        logger.info(f"Contact import: {result.message}")

        return result

    async def import_projects(
        self,
        content: bytes,
        filename: str | None = None,
        custom_mapping: dict[str, str] | None = None,
        update_existing: bool = True,
    ) -> ImportResult:
        """
        Import projects from file.

        Args:
            content: File content as bytes
            filename: Original filename for format detection
            custom_mapping: Custom column mapping to override defaults
            update_existing: Whether to update existing projects by ID

        Returns:
            ImportResult with counts and errors
        """
        format_type = _detect_format(content, filename)
        mapping = {**PROJECT_COLUMN_MAPPING, **(custom_mapping or {})}

        try:
            if format_type == "csv":
                raw_data = _parse_csv(content)
            elif format_type == "xlsx":
                raw_data = _parse_xlsx(content, sheet_name="Projets")
            else:
                raw_data = _parse_json(content)
        except Exception as e:
            return ImportResult(
                success=False,
                errors=[ImportError(row=0, column=None, message=str(e))],
            )

        result = ImportResult(success=True, total_rows=len(raw_data))

        # Status mapping
        status_map = {
            "en_cours": "active",
            "en cours": "active",
            "actif": "active",
            "active": "active",
            "en_pause": "on_hold",
            "en pause": "on_hold",
            "pause": "on_hold",
            "on_hold": "on_hold",
            "termine": "completed",
            "terminé": "completed",
            "completed": "completed",
            "annule": "cancelled",
            "annulé": "cancelled",
            "cancelled": "cancelled",
        }

        for idx, row in enumerate(raw_data):
            try:
                mapped = _map_columns(row, mapping)

                errors = _validate_project(mapped)
                if errors:
                    result.errors.append(ImportError(
                        row=idx + 1,
                        column=None,
                        message="; ".join(errors),
                        data=mapped,
                    ))
                    result.skipped += 1
                    continue

                project_id = mapped.get("id")
                existing = None

                if project_id:
                    stmt = select(Project).where(Project.id == project_id)
                    db_result = await self.session.execute(stmt)
                    existing = db_result.scalar_one_or_none()

                # Validate contact_id exists
                contact_id = mapped.get("contact_id")
                if contact_id:
                    stmt = select(Contact).where(Contact.id == contact_id)
                    db_result = await self.session.execute(stmt)
                    if not db_result.scalar_one_or_none():
                        contact_id = None  # Don't link to non-existent contact

                # Parse status
                raw_status = str(mapped.get("status", "active")).lower().strip()
                status = status_map.get(raw_status, "active")

                if existing and update_existing:
                    existing.name = mapped.get("name") or existing.name
                    if mapped.get("description"):
                        existing.description = mapped["description"]
                    existing.contact_id = contact_id
                    existing.status = status
                    if mapped.get("budget"):
                        existing.budget = _parse_value(mapped["budget"], "float")
                    if mapped.get("notes"):
                        existing.notes = mapped["notes"]
                    if mapped.get("tags"):
                        existing.tags = _parse_value(mapped["tags"], "tags")

                    existing.updated_at = datetime.now(UTC)
                    self.session.add(existing)
                    result.updated += 1

                elif existing and not update_existing:
                    result.skipped += 1

                else:
                    project = Project(
                        id=project_id or generate_uuid(),
                        name=mapped.get("name", "Sans nom"),
                        description=mapped.get("description"),
                        contact_id=contact_id,
                        status=status,
                        budget=_parse_value(mapped.get("budget"), "float"),
                        notes=mapped.get("notes"),
                        tags=_parse_value(mapped.get("tags"), "tags"),
                        scope="global",
                    )
                    self.session.add(project)
                    result.created += 1

            except Exception as e:
                logger.error(f"Error importing project row {idx + 1}: {e}")
                result.errors.append(ImportError(row=idx + 1, column=None, message=str(e)))
                result.skipped += 1

        await self.session.commit()

        result.success = len(result.errors) == 0
        logger.info(f"Project import: {result.message}")

        return result

    async def import_deliverables(
        self,
        content: bytes,
        filename: str | None = None,
        custom_mapping: dict[str, str] | None = None,
        update_existing: bool = True,
    ) -> ImportResult:
        """
        Import deliverables from file.

        Args:
            content: File content as bytes
            filename: Original filename for format detection
            custom_mapping: Custom column mapping to override defaults
            update_existing: Whether to update existing deliverables by ID

        Returns:
            ImportResult with counts and errors
        """
        format_type = _detect_format(content, filename)
        mapping = {**DELIVERABLE_COLUMN_MAPPING, **(custom_mapping or {})}

        try:
            if format_type == "csv":
                raw_data = _parse_csv(content)
            elif format_type == "xlsx":
                raw_data = _parse_xlsx(content, sheet_name="Livrables")
            else:
                raw_data = _parse_json(content)
        except Exception as e:
            return ImportResult(
                success=False,
                errors=[ImportError(row=0, column=None, message=str(e))],
            )

        result = ImportResult(success=True, total_rows=len(raw_data))

        # Status mapping
        status_map = {
            "a_faire": "a_faire",
            "a faire": "a_faire",
            "todo": "a_faire",
            "en_cours": "en_cours",
            "en cours": "en_cours",
            "in_progress": "en_cours",
            "en_revision": "en_revision",
            "en revision": "en_revision",
            "review": "en_revision",
            "valide": "valide",
            "validé": "valide",
            "done": "valide",
            "completed": "valide",
        }

        for idx, row in enumerate(raw_data):
            try:
                mapped = _map_columns(row, mapping)

                errors = _validate_deliverable(mapped)
                if errors:
                    result.errors.append(ImportError(
                        row=idx + 1,
                        column=None,
                        message="; ".join(errors),
                        data=mapped,
                    ))
                    result.skipped += 1
                    continue

                deliverable_id = mapped.get("id")
                existing = None

                if deliverable_id:
                    stmt = select(Deliverable).where(Deliverable.id == deliverable_id)
                    db_result = await self.session.execute(stmt)
                    existing = db_result.scalar_one_or_none()

                # Validate project_id exists
                project_id = mapped.get("project_id")
                if project_id:
                    stmt = select(Project).where(Project.id == project_id)
                    db_result = await self.session.execute(stmt)
                    if not db_result.scalar_one_or_none():
                        result.errors.append(ImportError(
                            row=idx + 1,
                            column="project_id",
                            message=f"Projet {project_id} non trouve",
                            data=mapped,
                        ))
                        result.skipped += 1
                        continue

                raw_status = str(mapped.get("status", "a_faire")).lower().strip()
                status = status_map.get(raw_status, "a_faire")

                if existing and update_existing:
                    existing.title = mapped.get("title") or existing.title
                    if mapped.get("description"):
                        existing.description = mapped["description"]
                    existing.project_id = project_id
                    existing.status = status
                    if mapped.get("due_date"):
                        existing.due_date = _parse_value(mapped["due_date"], "datetime")

                    existing.updated_at = datetime.now(UTC)
                    self.session.add(existing)
                    result.updated += 1

                elif existing and not update_existing:
                    result.skipped += 1

                else:
                    deliverable = Deliverable(
                        id=deliverable_id or generate_uuid(),
                        title=mapped.get("title", "Sans titre"),
                        description=mapped.get("description"),
                        project_id=project_id,
                        status=status,
                        due_date=_parse_value(mapped.get("due_date"), "datetime"),
                    )
                    self.session.add(deliverable)
                    result.created += 1

            except Exception as e:
                logger.error(f"Error importing deliverable row {idx + 1}: {e}")
                result.errors.append(ImportError(row=idx + 1, column=None, message=str(e)))
                result.skipped += 1

        await self.session.commit()

        result.success = len(result.errors) == 0
        logger.info(f"Deliverable import: {result.message}")

        return result
