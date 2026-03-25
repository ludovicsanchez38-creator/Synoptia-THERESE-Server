"""
THERESE v2 - Excel Generator Skill

Génère des fichiers Excel (.xlsx) avec le style Synoptia.
Approche code-execution avec fallback parser legacy.
"""

import json
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.services.skills.base import FileFormat, SkillParams, SkillResult
from app.services.skills.code_executor import CodeGenSkill

logger = logging.getLogger(__name__)


# Palette Synoptia pour Excel
SYNOPTIA_COLORS = {
    "header_bg": "0F1E6D",
    "header_text": "E6EDF7",
    "row_alt": "F5F7FA",
    "accent_cyan": "22D3EE",
    "accent_magenta": "E11D8D",
    "input_blue": "3B82F6",
    "formula_black": "1A1A2E",
    "link_green": "22C55E",
}


class XlsxSkill(CodeGenSkill):
    """
    Skill de génération de fichiers Excel.

    Crée des fichiers .xlsx professionnels avec le style Synoptia.
    Approche code-execution : le LLM génère du code openpyxl.
    Fallback automatique vers l'ancien parser Markdown/JSON.
    """

    skill_id = "xlsx-pro"
    name = "Tableur Excel Professionnel"
    description = "Génère un fichier Excel structuré avec le style Synoptia"
    output_format = FileFormat.XLSX

    def __init__(self, output_dir: Path):
        super().__init__(output_dir)

    def get_system_prompt_addition(self) -> str:
        """Instructions pour le LLM : générer du code Python openpyxl."""
        return """
## Instructions pour génération de fichier Excel

Tu dois générer un **bloc de code Python** complet utilisant la bibliothèque `openpyxl`.
Le code sera exécuté dans un environnement sandboxé avec les variables suivantes pré-injectées :
- `output_path` (str) : chemin où sauvegarder le fichier .xlsx
- `title` (str) : sujet demandé par l'utilisateur (utilise-le comme inspiration, mais génère un titre professionnel et pertinent, PAS le prompt brut)
- `SYNOPTIA_COLORS` (dict) : palette de couleurs Synoptia

### Imports disponibles
openpyxl (Workbook, Font, PatternFill, Alignment, Border, Side, BarChart, LineChart, PieChart, Reference, get_column_letter), datetime, json, re, math, Decimal

### Règles impératives
1. **Formules Excel** : utilise des formules Excel natives (=SUM, =AVERAGE, =IF, =VLOOKUP). Ne JAMAIS calculer en Python et écrire le résultat. Excel doit pouvoir recalculer.
2. **Multi-onglets** : crée plusieurs feuilles si le contenu le justifie (ex: Données, Résumé, Graphiques)
3. **Graphiques** : utilise openpyxl.chart (BarChart, LineChart, PieChart) pour visualiser les données quand c'est pertinent
4. **Formatage nombres** :
   - Monétaire : `cell.number_format = '#,##0.00 €'`
   - Pourcentages : `cell.number_format = '0.0%'`
   - Négatifs entre parenthèses : `cell.number_format = '#,##0.00;(#,##0.00)'`
5. **Couleurs financières** :
   - Bleu (inputs utilisateur) : `Font(color="3B82F6")`
   - Noir (formules calculées) : `Font(color="1A1A2E")`
   - Vert (liens entre feuilles) : `Font(color="22C55E")`
6. **Style header** : fond `PatternFill(start_color="0F1E6D", fill_type="solid")`, texte `Font(bold=True, color="E6EDF7")`
7. **Alternance lignes** : une ligne sur deux avec fond `PatternFill(start_color="F5F7FA", fill_type="solid")`
8. **Auto-fit colonnes** : ajuster la largeur des colonnes au contenu
9. **Footer** : dernière ligne = "Généré par THERESE - Synoptia" en italique gris
10. **Finir par** : `wb.save(output_path)`

**RÈGLE ABSOLUE** : Si les données demandées ne sont pas présentes dans le contexte fourni (contacts, projets, fichiers joints), NE PAS inventer de données. Générer à la place un modèle de tableau avec des en-têtes appropriés et des exemples de lignes clairement marquées comme "[Exemple - à remplacer]". Mentionner dans une cellule colorée : "Données non disponibles - modèle vierge fourni".

**INTERDIT** : Ne pas ajouter de bloc récapitulatif (Sujet / Action / Date) après le fichier. Le code se termine par `wb.save(output_path)`.

### Structure du code

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Données"

# ... construction du tableur ...

wb.save(output_path)
```

Génère UNIQUEMENT le bloc ```python``` avec le code complet. Pas d'explication avant ou après.
"""

    def get_markdown_prompt_addition(self) -> str:
        """Instructions Markdown pour modèles non code-capable."""
        return """
## Instructions pour génération de fichier Excel

Génère les données du tableur sous forme de tableaux Markdown.
Chaque tableau sera converti en un onglet Excel.
Commence chaque tableau par ## Nom de l'onglet.

### Format attendu
Utilise le format tableau Markdown standard :

## Données principales
| Header1 | Header2 | Header3 |
|---------|---------|---------|
| valeur1 | valeur2 | valeur3 |
| valeur4 | valeur5 | valeur6 |

### Règles
- Utilise des données réalistes et cohérentes
- Les nombres doivent être écrits sans formatage (ex: 1500, pas 1 500 €)
- Tu peux créer plusieurs tableaux (= plusieurs onglets)
- Ajoute un tableau de résumé si pertinent

NE génère PAS de code Python. Écris directement les tableaux de données.
"""

    async def _fallback_execute(
        self, params: SkillParams, file_id: str, output_path: Path
    ) -> SkillResult:
        """
        Fallback : ancien parser Markdown/JSON -> openpyxl.

        Args:
            params: Paramètres de génération
            file_id: ID du fichier pré-généré
            output_path: Chemin de sortie pré-calculé

        Returns:
            Résultat avec chemin vers le fichier généré
        """
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Données"

        # Parser le contenu et remplir le fichier
        data = self._parse_content(params.content)

        # Ajouter les données avec style
        self._add_data(ws, data, params.title)

        # Ajuster les largeurs de colonnes
        self._auto_fit_columns(ws)

        # Sauvegarder
        wb.save(str(output_path))

        # Calculer la taille
        file_size = output_path.stat().st_size

        logger.info(f"Generated XLSX (fallback): {output_path} ({file_size} bytes)")

        return SkillResult(
            file_id=file_id,
            file_path=output_path,
            file_name=output_path.name,
            file_size=file_size,
            mime_type=self.get_mime_type(),
            format=self.output_format,
        )

    def _parse_content(self, content: str) -> dict[str, Any]:
        """
        Parse le contenu en structure de données Excel.

        Args:
            content: Contenu généré par le LLM

        Returns:
            Dictionnaire avec headers, rows et formulas
        """
        # Essayer de parser comme JSON
        try:
            # Chercher un bloc JSON dans le contenu
            json_match = re.search(r'\{[\s\S]*"headers"[\s\S]*\}', content)
            if json_match:
                return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass

        # Sinon, parser comme tableau Markdown
        return self._parse_markdown_table(content)

    def _parse_markdown_table(self, content: str) -> dict[str, Any]:
        """
        Parse un tableau Markdown.

        Args:
            content: Contenu avec tableau Markdown

        Returns:
            Dictionnaire avec headers et rows
        """
        headers = []
        rows = []

        lines = content.strip().split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('|---') or re.match(r'^\|[\s\-:|]+\|$', line):
                continue

            if line.startswith('|'):
                cells = [c.strip() for c in line.split('|')[1:-1]]
                if cells:
                    if not headers:
                        headers = cells
                    else:
                        # Convertir les nombres
                        processed_cells = []
                        for cell in cells:
                            try:
                                if '.' in cell:
                                    processed_cells.append(float(cell))
                                else:
                                    processed_cells.append(int(cell))
                            except ValueError:
                                processed_cells.append(cell)
                        rows.append(processed_cells)

        return {
            "title": "Tableau",
            "headers": headers or ["A", "B", "C"],
            "rows": rows,
            "formulas": {},
        }

    def _add_data(self, ws, data: dict[str, Any], title: str) -> None:
        """
        Ajoute les données au worksheet avec le style Synoptia.

        Args:
            ws: Worksheet
            data: Données structurées
            title: Titre du document
        """
        # Styles
        header_fill = PatternFill(start_color=SYNOPTIA_COLORS["header_bg"],
                                  end_color=SYNOPTIA_COLORS["header_bg"],
                                  fill_type="solid")
        header_font = Font(name="Outfit", size=11, bold=True,
                          color=SYNOPTIA_COLORS["header_text"])
        data_font = Font(name="Inter", size=10)
        formula_font = Font(name="Inter", size=10, color=SYNOPTIA_COLORS["formula_black"])
        alt_fill = PatternFill(start_color=SYNOPTIA_COLORS["row_alt"],
                               end_color=SYNOPTIA_COLORS["row_alt"],
                               fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Titre du document (ligne 1)
        ws.merge_cells(start_row=1, start_column=1,
                      end_row=1, end_column=len(data.get("headers", ["A"])))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = data.get("title", title)
        title_cell.font = Font(name="Outfit", size=16, bold=True,
                              color=SYNOPTIA_COLORS["header_bg"])
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30

        # Ligne vide
        ws.row_dimensions[2].height = 10

        # Headers (ligne 3)
        headers = data.get("headers", [])
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[3].height = 25

        # Données (à partir de la ligne 4)
        rows = data.get("rows", [])
        for row_idx, row in enumerate(rows, start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border

                # Alignement selon le type
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = left_align

            # Alternance de couleurs
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        # Formules
        formulas = data.get("formulas", {})
        for cell_ref, formula in formulas.items():
            try:
                cell = ws[cell_ref]
                cell.value = formula
                cell.font = formula_font
                cell.alignment = Alignment(horizontal='right', vertical='center')
            except Exception as e:
                logger.warning(f"Could not add formula to {cell_ref}: {e}")

        # Footer Synoptia
        footer_row = len(rows) + 6
        ws.cell(row=footer_row, column=1).value = "Généré par THERESE - Synoptia"
        ws.cell(row=footer_row, column=1).font = Font(name="Inter", size=8, italic=True,
                                                      color="999999")

    def _auto_fit_columns(self, ws) -> None:
        """Ajuste automatiquement la largeur des colonnes."""
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass

            # Limiter entre 10 et 50 caractères
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
